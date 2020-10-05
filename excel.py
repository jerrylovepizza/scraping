import os
import gc
import logging
import argparse
import re
import openpyxl
import sys
import time
import csv

from openpyxl import Workbook
from pathlib import Path
from datetime import date

def checkValue(param):
    if (param is None):
        return ""
    
    if isinstance(param, date):
        return param.strftime("%m/%d/%Y")

    return str(param)

def read_excel():
    
    content = ""
    with open('sample.txt', 'r') as content_file:
        content = content_file.read()


    path = "data.xlsx"
    wb_obj = openpyxl.load_workbook(path) 
    sheet_obj = wb_obj.active 
    
    for i in range(2, 50):
        sample = content
        cell = sheet_obj.cell(i, 1)
        
        name = sheet_obj.cell(i, 1).value + " " + sheet_obj.cell(i, 2).value
        nationality = checkValue(sheet_obj.cell(i, 6).value )
        gender =  sheet_obj.cell(i, 4).value 
        if (gender.lower()[0] == "f"):
            gender = "Female"
        else:
            gender = "Male"
        birthday = sheet_obj.cell(i, 3).value        
        year = str(birthday.year)
        birthday = birthday.strftime("%m/%d/%Y")

        sample = sample.replace("$$name$$", name)
        sample = sample.replace("$$nationality$$", nationality)
        sample = sample.replace("$$gender$$", gender)
        sample = sample.replace("$$birthday$$", birthday)
        sample = sample.replace("$$year$$", year)

        height = sheet_obj.cell(i, 5).value
        address1 = sheet_obj.cell(i, 7).value
        address2 = checkValue(sheet_obj.cell(i, 8).value)
        city = checkValue(sheet_obj.cell(i, 9).value)
        state = checkValue(sheet_obj.cell(i, 10).value)
        zip = str(sheet_obj.cell(i, 11).value)
        country = checkValue( sheet_obj.cell(i, 12).value)
        email =checkValue( sheet_obj.cell(i, 13).value)
        parent1 =  checkValue(sheet_obj.cell(i, 18).value)
        parent1email = checkValue(sheet_obj.cell(i, 19).value)
        parent1tel = checkValue(sheet_obj.cell(i, 20).value)

        

        sample = sample.replace("$$height$$", height)
        sample = sample.replace("$$address1$$", address1)
        sample = sample.replace("$$address2$$", address2)
        sample = sample.replace("$$city$$", city)
        sample = sample.replace("$$zip$$", zip)
        sample = sample.replace("$$state$$", state)
        sample = sample.replace("$$country$$", country)
        sample = sample.replace("$$email$$", email)

        sample = sample.replace("$$parent1$$", parent1)
        sample = sample.replace("$$parent1email$$", parent1email)
        sample = sample.replace("$$parent1tel$$", parent1tel)



        parent2 =  checkValue(sheet_obj.cell(i, 21).value )       
        parent2email = checkValue(sheet_obj.cell(i, 22).value)
        parent2tel = checkValue(sheet_obj.cell(i, 23).value)

        sample = sample.replace("$$parent2$$", parent2)
        sample = sample.replace("$$parent2email$$", parent2email)
        sample = sample.replace("$$parent2tel$$", parent2tel)

        position =checkValue( sheet_obj.cell(i, 17).value)
        club = checkValue(sheet_obj.cell(i, 14).value)
        league = checkValue(sheet_obj.cell(i, 16).value)

        sample = sample.replace("$$position$$", position)
        sample = sample.replace("$$club$$", club)
        sample = sample.replace("$$league$$", league)


        status =checkValue( sheet_obj.cell(i, 25).value)
        highschool = checkValue(sheet_obj.cell(i, 24).value)
        graduation =  str(sheet_obj.cell(i, 26).value)

        sample = sample.replace("$$status$$", status)
        sample = sample.replace("$$highschool$$", highschool)
        sample = sample.replace("$$graduation$$", graduation)

        gpa = str(sheet_obj.cell(i, 27).value)
        major = checkValue(sheet_obj.cell(i, 30).value)
        act =  checkValue(sheet_obj.cell(i, 29).value)
        ncaa =  str(sheet_obj.cell(i, 31).value)
        naia = checkValue(sheet_obj.cell(i, 32).value)
        sat = checkValue(sheet_obj.cell(i, 28).value)
        
        sample = sample.replace("$$sat$$", sat)

        sample = sample.replace("$$gpa$$", gpa)
        sample = sample.replace("$$major$$", major)
        sample = sample.replace("$$ncaa$$", ncaa)
        sample = sample.replace("$$act$$", act)
        sample = sample.replace("$$naia$$", naia)

        
        filename = "res\\" + "{:02d}".format(i) + ".txt"
        with open(filename, 'w') as content_file:
            content_file.write(sample)
        # print(name, birthday, year,nationality, height, address1, address2, city, state, zip, country, email, parent1, parent1email, parent1tel, parent2, parent2email, parent2tel,position, club, league, status, highschool,graduation,  gpa, major, act, ncaa, naia)

        # print("\n")
read_excel()
